import csv
import json
import sqlite3
from typing import List, Dict, Any, Optional
import os
import hashlib


class DataRecorder:
    """
    数据记录器类，用于记录和保存爬虫获取的数据
    支持CSV、JSON、Markdown和SQLite格式的输出
    支持增量爬取功能
    """
    
    def __init__(self, output_file: str, primary_key: str = 'id', db_table: str = 'data'):
        """
        初始化数据记录器
        
        Args:
            output_file: 输出文件路径
            primary_key: 主键字段名，用于增量爬取时识别重复数据
            db_table: SQLite表名
        """
        self.output_file = output_file
        self.data = []
        self.primary_key = primary_key
        self.db_table = db_table
        # 用于存储已存在的数据主键值，用于增量爬取
        self.existing_keys = set()
        
        # 如果是SQLite数据库，检查并连接数据库
        if self.output_file.endswith('.db'):
            self._init_sqlite()
        # 如果是其他格式，尝试加载已存在的数据主键
        else:
            self._load_existing_keys()
    
    def add_data(self, data: List[Dict[str, Any]], incremental: bool = True):
        """
        添加数据到记录器，支持增量爬取
        
        Args:
            data: 要添加的数据列表或单个数据字典
            incremental: 是否启用增量模式，如果为True，则只添加新数据
        """
        if not data:
            return
        
        # 确保数据是列表格式
        data_list = data if isinstance(data, list) else [data]
        
        if incremental and self.existing_keys:
            # 增量模式：只添加新数据
            new_data = []
            for item in data_list:
                # 获取主键值，如果不存在则生成一个基于数据的哈希值
                key_value = item.get(self.primary_key)
                if key_value is None:
                    # 如果没有主键，创建一个基于数据内容的哈希值
                    key_value = self._generate_data_hash(item)
                
                # 只添加主键不存在的数据
                if key_value not in self.existing_keys:
                    new_data.append(item)
                    # 将新数据的主键添加到已存在的键集合中
                    self.existing_keys.add(key_value)
            
            if new_data:
                self.data.extend(new_data)
                print(f"增量添加了 {len(new_data)} 条新数据")
            else:
                print("没有新数据需要添加")
        else:
            # 非增量模式：添加所有数据
            self.data.extend(data_list)
            print(f"添加了 {len(data_list)} 条数据")
            
            # 更新已存在的键集合
            for item in data_list:
                key_value = item.get(self.primary_key)
                if key_value is None:
                    key_value = self._generate_data_hash(item)
                self.existing_keys.add(key_value)
    
    def clear_data(self):
        """
        清空记录器中的数据
        """
        self.data = []
        
    def _init_sqlite(self):
        """
        初始化SQLite数据库连接
        """
        # 检查数据库文件是否存在
        db_exists = os.path.exists(self.output_file)
        
        # 连接数据库
        self.conn = sqlite3.connect(self.output_file)
        self.cursor = self.conn.cursor()
        
        # 如果数据库不存在，创建表结构
        if not db_exists:
            # 暂时不创建表，等有数据时根据数据结构创建
            print(f"创建了新的SQLite数据库: {self.output_file}")
        else:
            # 检查表是否存在
            self.cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{self.db_table}'")
            if self.cursor.fetchone():
                # 加载已存在的数据主键
                try:
                    self.cursor.execute(f"SELECT {self.primary_key} FROM {self.db_table}")
                    for row in self.cursor.fetchall():
                        if row[0] is not None:
                            self.existing_keys.add(row[0])
                    print(f"从数据库加载了 {len(self.existing_keys)} 条已存在的数据记录")
                except sqlite3.OperationalError:
                    # 如果主键列不存在，忽略
                    print(f"数据库表 {self.db_table} 中不存在主键列 {self.primary_key}")
            else:
                print(f"数据库中不存在表 {self.db_table}，将在保存数据时创建")
    
    def _load_existing_keys(self):
        """
        从已存在的文件中加载数据主键
        """
        if not os.path.exists(self.output_file):
            return
        
        try:
            if self.output_file.endswith('.json'):
                with open(self.output_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                    for item in existing_data:
                        key_value = item.get(self.primary_key)
                        if key_value is not None:
                            self.existing_keys.add(key_value)
            elif self.output_file.endswith('.csv'):
                with open(self.output_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if self.primary_key in row:
                            key_value = row[self.primary_key]
                            if key_value:
                                self.existing_keys.add(key_value)
            
            if self.existing_keys:
                print(f"从现有文件加载了 {len(self.existing_keys)} 条数据记录")
        except Exception as e:
            print(f"加载现有数据键时出错: {str(e)}")
    
    def _generate_data_hash(self, data: Dict[str, Any]) -> str:
        """
        基于数据内容生成哈希值，作为没有主键时的唯一标识
        
        Args:
            data: 数据字典
            
        Returns:
            str: 哈希值
        """
        # 将字典转换为排序后的字符串
        sorted_items = sorted(data.items())
        data_str = str(sorted_items)
        # 生成MD5哈希
        return hashlib.md5(data_str.encode('utf-8')).hexdigest()
    
    def record(self, file_format: Optional[str] = None, remove_columns: Optional[List[str]] = None):
        """
        记录数据到文件
        
        Args:
            file_format: 文件格式，如果为None则根据文件扩展名自动判断
            remove_columns: 需要移除的列名列表（适用于Markdown格式）
        """
        if not self.data:
            print("没有数据需要记录")
            return
        
        # 如果没有指定格式，则根据文件扩展名判断
        if file_format is None:
            if self.output_file.endswith('.db'):
                file_format = 'sqlite'
            elif self.output_file.endswith('.csv'):
                file_format = 'csv'
            elif self.output_file.endswith('.json'):
                file_format = 'json'
            elif self.output_file.endswith('.md'):
                file_format = 'markdown'
            elif self.output_file.endswith('.xlsx') or self.output_file.endswith('.xls'):
                file_format = 'excel'
            else:
                print(f"无法识别文件格式，默认使用JSON格式: {self.output_file}")
                file_format = 'json'
        
        # 根据格式调用相应的记录方法
        if file_format.lower() == 'csv':
            self._record_to_csv()
        elif file_format.lower() == 'json':
            self._record_to_json()
        elif file_format.lower() == 'markdown':
            self._record_to_markdown(remove_columns)
        elif file_format.lower() == 'sqlite':
            self._record_to_sqlite()
        elif file_format.lower() == 'excel':
            self._record_to_excel()
        else:
            print(f"不支持的文件格式: {file_format}")
    
    def _record_to_csv(self):
        """
        将数据记录为CSV格式
        支持增量模式：如果文件已存在，则追加数据
        """
        if not self.data:
            return
        
        # 获取所有字段名
        fieldnames = set()
        for item in self.data:
            fieldnames.update(item.keys())
        fieldnames = sorted(fieldnames)  # 排序以保持一致性
        
        # 检查文件是否已存在
        file_exists = os.path.exists(self.output_file)
        
        try:
            # 如果是增量添加且文件存在，尝试以追加模式打开
            if file_exists and hasattr(self, 'existing_keys') and self.existing_keys:
                # 在增量模式下，我们仍然重写整个文件，但只包含新数据
                # 这样可以确保文件格式的一致性
                with open(self.output_file, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    # 先写入已存在的数据（如果能够加载）
                    existing_data = []
                    try:
                        if os.path.exists(self.output_file):
                            with open(self.output_file, 'r', encoding='utf-8') as f:
                                reader = csv.DictReader(f)
                                existing_data = list(reader)
                    except Exception:
                        pass
                    
                    # 写入现有数据
                    for item in existing_data:
                        writer.writerow(item)
                    
                    # 写入新数据
                    for item in self.data:
                        row = {field: item.get(field, '') for field in fieldnames}
                        writer.writerow(row)
            else:
                # 正常写入模式
                with open(self.output_file, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for item in self.data:
                        row = {field: item.get(field, '') for field in fieldnames}
                        writer.writerow(row)
            print(f"数据已保存到CSV文件: {self.output_file}")
        except Exception as e:
            print(f"保存CSV文件时出错: {str(e)}")
    
    def _record_to_json(self):
        """
        将数据记录为JSON格式
        支持增量模式：如果文件已存在，则合并数据
        """
        if not self.data:
            return
        
        # 检查是否是增量模式且文件存在
        file_exists = os.path.exists(self.output_file)
        final_data = self.data.copy()
        
        # 如果是增量模式且文件存在，尝试合并数据
        if file_exists and hasattr(self, 'existing_keys') and self.existing_keys:
            try:
                with open(self.output_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                    # 合并数据
                    final_data = existing_data + self.data
            except Exception:
                # 如果读取失败，使用新数据
                pass
        
        try:
            with open(self.output_file, 'w', encoding='utf-8') as jsonfile:
                json.dump(final_data, jsonfile, ensure_ascii=False, indent=2)
            print(f"数据已保存到JSON文件: {self.output_file}")
        except Exception as e:
            print(f"保存JSON文件时出错: {str(e)}")
    
    def _record_to_markdown(self, remove_columns: Optional[List[str]] = None):
        """
        将数据记录为Markdown表格格式
        
        Args:
            remove_columns: 需要移除的列名列表
        """
        if not self.data:
            return
        
        # 处理移除列参数
        remove_columns = remove_columns or []
        
        # 获取所有字段名并过滤
        fieldnames = set()
        for item in self.data:
            fieldnames.update(item.keys())
        
        # 过滤掉需要移除的列
        headers = [header for header in fieldnames if header not in remove_columns]
        headers = sorted(headers)  # 排序表头，保持一致性
        
        # 构建Markdown表格内容
        md_lines = []
        # 表头行
        md_lines.append("| " + " | ".join(headers) + " |")
        # 分隔线
        md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # 添加每行数据（只包含过滤后的列）
        for item in self.data:
            row = [str(item.get(header, "")) for header in headers]
            md_lines.append("| " + " | ".join(row) + " |")
        
        try:
            with open(self.output_file, 'w', encoding='utf-8') as mdfile:
                mdfile.write("\n".join(md_lines))
            print(f"数据已保存到Markdown文件: {self.output_file}")
        except Exception as e:
            print(f"保存Markdown文件时出错: {str(e)}")
    
    def _record_to_sqlite(self):
        """
        将数据记录到SQLite数据库
        """
        if not self.data:
            return
        
        try:
            # 确保数据库连接存在
            if not hasattr(self, 'conn'):
                self._init_sqlite()
            
            # 获取所有字段名
            fieldnames = set()
            for item in self.data:
                fieldnames.update(item.keys())
            fieldnames = sorted(fieldnames)  # 排序以保持一致性
            
            # 检查表是否存在，如果不存在则创建
            self.cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{self.db_table}'")
            if not self.cursor.fetchone():
                # 创建表结构
                # 注意：这里简化处理，所有字段都设为TEXT类型
                columns_def = ", ".join([f"{field} TEXT" for field in fieldnames])
                # 如果主键字段存在，将其设为主键
                if self.primary_key in fieldnames:
                    # 先移除主键，再添加为主键
                    columns_def = ", ".join([f"{field} TEXT PRIMARY KEY" if field == self.primary_key else f"{field} TEXT" for field in fieldnames])
                
                create_table_sql = f"CREATE TABLE {self.db_table} ({columns_def})"
                self.cursor.execute(create_table_sql)
                print(f"创建了数据库表: {self.db_table}")
            
            # 插入数据
            for item in self.data:
                # 准备插入的数据
                values = [item.get(field, '') for field in fieldnames]
                placeholders = ", ".join(["?"] * len(fieldnames))
                
                # 如果存在主键，使用INSERT OR REPLACE
                if self.primary_key in fieldnames:
                    insert_sql = f"INSERT OR REPLACE INTO {self.db_table} ({', '.join(fieldnames)}) VALUES ({placeholders})"
                else:
                    insert_sql = f"INSERT INTO {self.db_table} ({', '.join(fieldnames)}) VALUES ({placeholders})"
                
                try:
                    self.cursor.execute(insert_sql, values)
                except sqlite3.IntegrityError:
                    # 忽略主键冲突
                    pass
            
            # 提交事务
            self.conn.commit()
            print(f"数据已保存到SQLite数据库: {self.output_file}，表名: {self.db_table}")
        except Exception as e:
            print(f"保存SQLite数据库时出错: {str(e)}")
            # 回滚事务
            if hasattr(self, 'conn'):
                self.conn.rollback()
    
    def _record_to_excel(self):
        """
        将数据记录为Excel格式
        """
        if not self.data:
            return
        
        try:
            # 尝试导入openpyxl库
            import openpyxl
            
            # 获取所有字段名
            fieldnames = set()
            for item in self.data:
                fieldnames.update(item.keys())
            fieldnames = sorted(fieldnames)  # 排序以保持一致性
            
            # 检查文件是否存在
            file_exists = os.path.exists(self.output_file)
            
            if file_exists:
                # 打开现有文件
                workbook = openpyxl.load_workbook(self.output_file)
                # 使用第一个工作表
                if self.db_table in workbook.sheetnames:
                    sheet = workbook[self.db_table]
                else:
                    sheet = workbook.active
                    sheet.title = self.db_table
                
                # 检查是否有表头
                has_header = False
                if sheet.max_row > 0:
                    # 假设第一行是表头
                    has_header = True
                
                # 如果有增量数据，直接追加
                start_row = sheet.max_row + 1
            else:
                # 创建新工作簿
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = self.db_table
                start_row = 1
                has_header = False
            
            # 写入表头（如果需要）
            if not has_header:
                for col_idx, field in enumerate(fieldnames, 1):
                    sheet.cell(row=start_row, column=col_idx, value=field)
                start_row += 1
            
            # 写入数据
            for item in self.data:
                for col_idx, field in enumerate(fieldnames, 1):
                    value = item.get(field, '')
                    sheet.cell(row=start_row, column=col_idx, value=value)
                start_row += 1
            
            # 保存文件
            workbook.save(self.output_file)
            print(f"数据已保存到Excel文件: {self.output_file}，工作表名: {self.db_table}")
        except ImportError:
            print("需要安装openpyxl库来支持Excel格式。请运行: pip install openpyxl")
        except Exception as e:
            print(f"保存Excel文件时出错: {str(e)}")
    
    def close(self):
        """
        关闭数据库连接等资源
        """
        if hasattr(self, 'conn'):
            self.conn.close()
            print("数据库连接已关闭")
    
    def __del__(self):
        """
        析构函数，确保资源被释放
        """
        self.close()
    
    def record_to_markdown(self, md_file_path: str, remove_columns: Optional[List[str]] = None):
        """
        将数据记录到指定的Markdown文件
        
        Args:
            md_file_path: Markdown文件路径
            remove_columns: 需要移除的列名列表
        """
        # 保存当前的输出文件路径
        original_output = self.output_file
        # 临时修改输出文件路径
        self.output_file = md_file_path
        # 调用内部方法记录数据
        self._record_to_markdown(remove_columns)
        # 恢复原始输出文件路径
        self.output_file = original_output
    
    def get_data_count(self) -> int:
        """
        获取当前记录器中的数据数量
        
        Returns:
            int: 数据数量
        """
        return len(self.data)
    
    def get_sample_data(self, count: int = 1) -> List[Dict[str, Any]]:
        """
        获取样本数据
        
        Args:
            count: 样本数量
            
        Returns:
            List[Dict[str, Any]]: 样本数据列表
        """
        return self.data[:count]